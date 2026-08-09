"""Spreadsheet load / validate, plus a per-file log of completed entries.

Model: the original download is never modified. As you label, every *completed*
entry is written live into ``<original name>_labeled.<ext>`` inside the output
folder, keyed by ML catalog number. Each Macaulay download is one species, so the
output folder accumulates one labeled file per species over time. Reopening a
file restores its prior labels so you resume where you left off.

Each write is a full atomic rewrite (temp file + rename), so a crash mid-write
can't corrupt the file. Input and output may be CSV or XLSX.
"""

from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path

# Minimum columns we rely on. Missing any of these is a hard error.
REQUIRED_COLUMNS = [
    "ML Catalog Number",
    "Format",
    "Common Name",
    "Scientific Name",
    "Asset Tags",
]
CATALOG_KEY = "ML Catalog Number"

# Columns bird-swipe appends. Order preserved when new to a file.
LABEL_COLUMNS = ["nest_label", "human_structure", "reviewed", "reviewed_at", "reviewer"]

REVIEWED = "TRUE"


def labeled_name(input_path: str | Path) -> str:
    """`ML__..._rethaw.csv` -> `ML__..._rethaw_labeled.csv` (extension kept)."""
    p = Path(input_path)
    return f"{p.stem}_labeled{p.suffix or '.csv'}"


class ValidationError(Exception):
    """Raised when a file doesn't look like a Macaulay export."""


class Validation:
    def __init__(self, errors: list[str], warnings: list[str]):
        self.errors = errors
        self.warnings = warnings

    @property
    def ok(self) -> bool:
        return not self.errors


def validate_fieldnames(fieldnames: list[str] | None) -> Validation:
    errors: list[str] = []
    warnings: list[str] = []
    if not fieldnames:
        return Validation(["File has no header row."], [])
    if fieldnames[0] != CATALOG_KEY:
        errors.append(f"First column must be {CATALOG_KEY!r}, got {fieldnames[0]!r}.")
    for col in REQUIRED_COLUMNS:
        if col not in fieldnames:
            errors.append(f"Missing required column: {col!r}.")
    # A genuine Macaulay export is wide; warn (don't fail) if it looks trimmed.
    data_cols = [c for c in fieldnames if c not in LABEL_COLUMNS]
    if len(data_cols) < 10:
        warnings.append(
            f"Only {len(data_cols)} data columns found; expected a full Macaulay export (~46)."
        )
    return Validation(errors, warnings)


def _is_xlsx(path: Path) -> bool:
    return path.suffix.lower() in (".xlsx", ".xlsm")


def _read_csv(path: Path) -> tuple[list[dict], list[str]]:
    # utf-8-sig strips a BOM if Macaulay/Excel added one.
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = [dict(r) for r in reader]
        fieldnames = list(reader.fieldnames or [])
    return rows, fieldnames


def _read_xlsx(path: Path) -> tuple[list[dict], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if header is None:
            return [], []
        fieldnames = [str(h) if h is not None else "" for h in header]
        rows = []
        for r in it:
            if r is None or all(c is None for c in r):
                continue  # skip blank rows
            rows.append({name: ("" if i >= len(r) or r[i] is None else str(r[i]))
                         for i, name in enumerate(fieldnames)})
        return rows, fieldnames
    finally:
        wb.close()


def _read_table(path: Path) -> tuple[list[dict], list[str]]:
    return _read_xlsx(path) if _is_xlsx(path) else _read_csv(path)


class LabeledFile:
    """One ``<name>_labeled.<ext>`` holding the completed entries for a species file.

    Rows are keyed by ML catalog number; re-labeling an asset updates its row.
    Written as CSV or XLSX to match the path's extension.
    """

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.rows_by_id: dict[str, dict] = {}
        self.fieldnames: list[str] = []
        self._load()

    def _load(self) -> None:
        if not self.path.exists():
            return
        rows, fieldnames = _read_table(self.path)
        self.fieldnames = fieldnames
        for r in rows:
            key = r.get(CATALOG_KEY)
            if key:
                self.rows_by_id[str(key)] = r

    def get(self, ml_id: str | int) -> dict | None:
        return self.rows_by_id.get(str(ml_id))

    def upsert(self, row: dict) -> None:
        """Insert or update a completed entry, then persist."""
        for col in row:  # grow the column union, preserving order
            if col not in self.fieldnames:
                self.fieldnames.append(col)
        self.rows_by_id[str(row[CATALOG_KEY])] = dict(row)
        self.save()

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        tmp = self.path.with_suffix(self.path.suffix + ".part")
        if _is_xlsx(self.path):
            self._write_xlsx(tmp)
        else:
            self._write_csv(tmp)
        tmp.replace(self.path)  # atomic on the same filesystem

    def _write_csv(self, path: Path) -> None:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=self.fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(self.rows_by_id.values())

    def _write_xlsx(self, path: Path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(self.fieldnames)
        for row in self.rows_by_id.values():
            ws.append([row.get(f, "") for f in self.fieldnames])
        wb.save(path)

    def count(self) -> int:
        return len(self.rows_by_id)


class Catalog:
    """Rows of one Macaulay export, with completed labels persisted to a LabeledFile."""

    def __init__(self, rows: list[dict], fieldnames: list[str], labeled: LabeledFile):
        self.rows = rows
        self.fieldnames = fieldnames
        self.labeled = labeled
        self._ensure_label_columns()

    # --- construction -----------------------------------------------------
    @classmethod
    def open(
        cls,
        input_path: str | Path,
        output_dir: str | Path,
        *,
        resume: bool = True,
    ) -> tuple["Catalog", Validation]:
        """Load ``input_path``; labels go to ``output_dir/<name>_labeled.<ext>``."""
        input_path = Path(input_path)
        rows, fieldnames = _read_table(input_path)
        validation = validate_fieldnames(fieldnames)
        if not validation.ok:
            raise ValidationError("; ".join(validation.errors))

        labeled = LabeledFile(Path(output_dir) / labeled_name(input_path))
        cat = cls(rows, fieldnames, labeled)
        if resume:
            cat._restore_from_labeled()
        return cat, validation

    def _ensure_label_columns(self) -> None:
        for col in LABEL_COLUMNS:
            if col not in self.fieldnames:
                self.fieldnames.append(col)
        for row in self.rows:
            for col in LABEL_COLUMNS:
                row.setdefault(col, "")

    def _restore_from_labeled(self) -> None:
        """Copy label fields back onto rows already present in the labeled file."""
        for row in self.rows:
            prior = self.labeled.get(row.get(CATALOG_KEY, ""))
            if prior and prior.get("reviewed") == REVIEWED:
                for col in LABEL_COLUMNS:
                    row[col] = prior.get(col, "")

    def retarget_output_dir(self, folder: str | Path) -> None:
        """Move the labeled file to a new folder, carrying completed rows across."""
        self.labeled = LabeledFile(Path(folder) / self.labeled.path.name)
        for row in self.rows:
            if row.get("reviewed") == REVIEWED:
                self.labeled.upsert(row)

    # --- labeling ---------------------------------------------------------
    def set_label(
        self, index: int, nest: bool | None, structure: bool, reviewer: str = ""
    ) -> None:
        """Record a decision for row ``index`` and persist it to the labeled file."""
        row = self.rows[index]
        row["nest_label"] = "" if nest is None else ("yes" if nest else "no")
        row["human_structure"] = "yes" if structure else "no"
        row["reviewed"] = REVIEWED
        row["reviewed_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
        row["reviewer"] = reviewer
        self.labeled.upsert(row)

    def is_reviewed(self, index: int) -> bool:
        return self.rows[index].get("reviewed") == REVIEWED

    def first_unreviewed(self) -> int:
        for i in range(len(self.rows)):
            if not self.is_reviewed(i):
                return i
        return len(self.rows)  # all done

    def stats(self) -> dict:
        reviewed = sum(1 for r in self.rows if r.get("reviewed") == REVIEWED)
        yes = sum(1 for r in self.rows if r.get("nest_label") == "yes")
        no = sum(1 for r in self.rows if r.get("nest_label") == "no")
        struct = sum(1 for r in self.rows if r.get("human_structure") == "yes")
        return {"total": len(self.rows), "reviewed": reviewed, "yes": yes, "no": no, "structure": struct}
